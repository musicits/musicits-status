#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""엑셀 서식 — 모든 도구의 xlsx가 같은 모양을 갖도록.

통합 전에는 이 헬퍼들이 7곳에 복붙돼 있었고 시그니처까지 어긋나 있었다
(7.경쟁벤치마킹만 cfont(b=...), 5·6번은 cfont(bold=...)). 설명서에 "헷갈리지
말 것"이라고 적어둬야 할 정도였다. 여기 하나로 합치면서 bold= 로 통일했다.

openpyxl 이 없으면 HAVE_OPENPYXL 이 False다. 도구 쪽에서 이걸 보고 조용히
건너뛴다 — 엑셀이 없다고 리포트.txt 까지 못 쓰게 만들 이유는 없다.
"""

try:
    from openpyxl import Workbook                       # noqa: F401
    from openpyxl.styles import (Alignment, Border, Font,   # noqa: F401
                                 PatternFill, Side)
    from openpyxl.utils import get_column_letter         # noqa: F401
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

NAVY = "1F3864"
BLUE = "2E5496"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
GREY = "808080"
LIGHT = "F2F2F2"

if HAVE_OPENPYXL:
    _thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
else:
    BORDER = None


def hfont(size=11, color="FFFFFF", bold=True):
    """머리글용 글꼴."""
    return Font(name="맑은 고딕", size=size, color=color, bold=bold)


def cfont(size=10, color="000000", bold=False):
    """본문 칸용 글꼴."""
    return Font(name="맑은 고딕", size=size, color=color, bold=bold)


def fill(color):
    return PatternFill("solid", fgColor=color)


def score_fill(score):
    """점수대별 배경색. 80↑ 초록 / 60↑ 노랑 / 그 아래 빨강."""
    if score == "" or score is None:
        return fill(LIGHT)
    if score >= 80:
        return fill(GREEN)
    if score >= 60:
        return fill(YELLOW)
    return fill(RED)


def title_block(ws, text, sub, ncols):
    """시트 맨 위 제목 두 줄(큰 제목 + 회색 설명)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = hfont(15)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, sub)
    c.font = cfont(9, GREY)
    c.fill = fill(LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def header_row(ws, row, headers, widths):
    """표 머리글 한 줄 + 열 너비."""
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row, i, h)
        c.font = hfont(10)
        c.fill = fill(BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def cell(ws, row, col, value, bold=False, color="000000", bg=None,
         align="left", wrap=False, size=10):
    """본문 칸 하나. 서식까지 한 번에."""
    c = ws.cell(row, col, value)
    c.font = cfont(size, color, bold=bold)
    if bg:
        c.fill = fill(bg) if isinstance(bg, str) else bg
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER
    return c


def notes_sheet(wb, title, lines, width=110):
    """'읽는법' 시트. 모든 도구가 같은 모양으로 붙인다."""
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = width
    title_block(ws, title, "이 표를 어떻게 읽는지 — 헷갈리면 여기부터", 1)
    row = 4
    for line in lines:
        c = ws.cell(row, 1, line)
        # 빈 줄이 아니고 마침표 없이 짧으면 소제목으로 본다
        head = bool(line) and not line.startswith(" ") and len(line) < 40 \
            and not line.endswith(".")
        c.font = cfont(11 if head else 10, NAVY if head else "000000", bold=head)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1
    return ws
